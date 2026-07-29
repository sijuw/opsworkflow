from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def generate_excel(samples, filename: str) -> str:
    """Write samples to generated_reports/<filename> and return the path.

    The caller supplies the filename so the name shown in the preview and
    the name that reaches the recipient are the same string.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sample Transactions"

    if samples:
        # Header
        headers = list(samples[0].keys())
        sheet.append(headers)

        # Make header bold
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        # Data
        for sample in samples:
            sheet.append(list(sample.values()))

        # Auto-size columns
        for column_cells in sheet.columns:
            max_length = 0
            column = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                try:
                    value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(value))
                except Exception:
                    pass

            sheet.column_dimensions[column].width = min(max_length + 2, 50)

    # Save inside a dedicated folder
    output_dir = Path("generated_reports")
    output_dir.mkdir(exist_ok=True)

    # .name strips any directory component an institution name could smuggle in
    filepath = output_dir / Path(filename).name
    workbook.save(filepath)

    return str(filepath)